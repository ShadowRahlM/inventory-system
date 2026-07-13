import json
import io

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
from django.db.models import Count, F, Q, Sum
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import Tile, Batch, Inventory, Movement, AuditLog, TileCatalog, Customer, Supplier, SalesOrder, PurchaseOrder, OrderLineItem, OrderStatus, Notification, SyncConflict
from .serializers import (
    TileSerializer, BatchSerializer, InventorySerializer, MovementSerializer, AuditLogSerializer, TileCatalogSerializer, ProcessCatalogSerializer,
    ReceiveInventorySerializer, DispatchInventorySerializer, AdjustInventorySerializer, TransferInventorySerializer,
    UserSerializer, SetRoleSerializer,
    CustomerSerializer, SupplierSerializer, SalesOrderSerializer, PurchaseOrderSerializer, OrderLineItemSerializer,
    CreateSalesOrderSerializer, CreatePurchaseOrderSerializer, ConfirmOrderSerializer,
    NotificationSerializer, SyncConflictSerializer, StockTakeSerializer,
)
from .services import InventoryService, OrderService
from .permissions import IsInventoryViewer, CanPerformInventoryOperations, IsAdminUser


def _extract_catalog_items(json_data):
    """Extract a flat list of tile item dicts from various catalog JSON formats.

    Supported formats:
    - Flat array: [{"sku": "...", "name": "...", ...}, ...]
    - image_to_sku_mappings: {"image_to_sku_mappings": [{"items": [{"sku_code": "...", "size": "...", ...}]}]}
    """
    if isinstance(json_data, list):
        return json_data
    if not isinstance(json_data, dict):
        return []
    mappings = json_data.get('image_to_sku_mappings')
    if isinstance(mappings, list):
        items = []
        for mapping in mappings:
            for entry in mapping.get('items', []):
                items.append({
                    'sku': entry.get('sku_code', ''),
                    'name': entry.get('sku_code', ''),
                    'dimensions': entry.get('size', '30x30cm'),
                    'category': entry.get('category', 'Wall'),
                    'pieces_per_carton': entry.get('pieces_per_carton', 10),
                })
        return items
    return []


class TileViewSet(viewsets.ModelViewSet):
    queryset = Tile.objects.all()
    serializer_class = TileSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['category', 'sku']
    search_fields = ['sku', 'name', 'description', 'category', 'brand', 'series', 'tile_type', 'finish', 'use_case']
    ordering_fields = ['sku', 'name', 'created_at']
    ordering = ['sku']

    @action(detail=False, methods=['get'])
    def check_sku(self, request):
        sku = request.query_params.get('sku', '')
        if not sku or len(sku) < 2:
            return Response({'exists': False, 'tile': None})
        try:
            tile = Tile.objects.get(sku__iexact=sku)
            return Response({'exists': True, 'tile': TileSerializer(tile).data})
        except Tile.DoesNotExist:
            return Response({'exists': False, 'tile': None})

    @action(detail=False, methods=['post'])
    def check_skus(self, request):
        skus = request.data.get('skus', [])
        if not isinstance(skus, list):
            return Response({'error': 'Expected a list of SKUs'}, status=400)
        existing = {}
        for sku in skus:
            if not isinstance(sku, str) or not sku.strip():
                continue
            try:
                tile = Tile.objects.get(sku__iexact=sku.strip())
                existing[sku] = TileSerializer(tile).data
            except Tile.DoesNotExist:
                pass
        return Response({'existing': existing})

    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'success': False, 'error': 'No IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            AuditLog.objects.filter(movement__tile_id__in=ids).delete()
            Movement.objects.filter(tile_id__in=ids).delete()
            Inventory.objects.filter(tile_id__in=ids).delete()
            Batch.objects.filter(tile_id__in=ids).delete()
            deleted = Tile.objects.filter(id__in=ids).delete()
        return Response({'success': True, 'data': {'deleted': deleted[0]}})


class BatchViewSet(viewsets.ModelViewSet):
    queryset = Batch.objects.select_related('tile').all()
    serializer_class = BatchSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['tile', 'is_active']
    search_fields = ['batch_number', 'supplier']
    ordering_fields = ['production_date', 'received_date']
    ordering = ['-received_date']


class InventoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Inventory.objects.select_related('tile', 'batch').all()
    serializer_class = InventorySerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['tile', 'batch', 'location']
    search_fields = ['location']
    ordering_fields = ['updated_at', 'location']
    ordering = ['-updated_at']

    @action(detail=False, methods=['get'])
    def available_stock(self, request):
        tile_id = request.query_params.get('tile_id')
        if not tile_id:
            return Response({'error': 'tile_id parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            stock_info = InventoryService.get_available_stock(tile_id)
            return Response(stock_info)
        except Tile.DoesNotExist:
            return Response({'error': 'Tile not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        threshold = int(request.query_params.get('threshold', 50))
        low = Inventory.objects.select_related('tile', 'batch').annotate(
            computed_total=F('cartons') * F('tile__pieces_per_carton') + F('loose_pieces')
        ).filter(computed_total__lte=threshold).order_by('computed_total')
        page = self.paginate_queryset(low)
        if page is not None:
            serializer = InventorySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = InventorySerializer(low, many=True)
        return Response({
            'count': len(serializer.data),
            'threshold': threshold,
            'results': serializer.data,
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory.csv"'
        writer = csv.writer(response)
        writer.writerow(['Tile SKU', 'Tile Name', 'Batch', 'Location', 'Cartons', 'Loose Pieces', 'Total Pieces'])
        for item in self.get_queryset():
            writer.writerow([
                item.tile.sku, item.tile.name, item.batch.batch_number,
                item.location, item.cartons, item.loose_pieces, item.total_pieces,
            ])
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from django.http import HttpResponse
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        headers = ['Tile SKU', 'Tile Name', 'Batch', 'Location', 'Cartons', 'Loose Pieces', 'Total Pieces']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        for row, item in enumerate(self.get_queryset(), 2):
            ws.cell(row=row, column=1, value=item.tile.sku)
            ws.cell(row=row, column=2, value=item.tile.name)
            ws.cell(row=row, column=3, value=item.batch.batch_number)
            ws.cell(row=row, column=4, value=item.location)
            ws.cell(row=row, column=5, value=item.cartons)
            ws.cell(row=row, column=6, value=item.loose_pieces)
            ws.cell(row=row, column=7, value=item.total_pieces)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventory.xlsx"'
        wb.save(response)
        return response


class MovementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Movement.objects.select_related('tile', 'batch', 'performed_by').all()
    serializer_class = MovementSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['tile', 'batch', 'movement_type']
    search_fields = ['reference', 'reason']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="movements.csv"'
        writer = csv.writer(response)
        writer.writerow(['Date', 'Type', 'Tile SKU', 'Batch', 'Cartons Change', 'Loose Change', 'Reference', 'Reason', 'Performed By'])
        for m in self.get_queryset():
            writer.writerow([
                m.created_at.isoformat(), m.movement_type, m.tile.sku, m.batch.batch_number,
                m.cartons_change, m.loose_pieces_change, m.reference, m.reason,
                m.performed_by.username,
            ])
        return response


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related('movement', 'changed_by').all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['movement', 'action']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']


class TileCatalogViewSet(viewsets.ModelViewSet):
    queryset = TileCatalog.objects.all()
    serializer_class = TileCatalogSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'success': False, 'error': 'No IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        deleted = TileCatalog.objects.filter(id__in=ids).delete()
        return Response({'success': True, 'data': {'deleted': deleted[0]}})

    @action(detail=True, methods=['post'])
    def process(self, request, pk=None):
        catalog = self.get_object()
        if catalog.processed:
            return Response(
                {'success': False, 'error': 'Catalog already processed'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        json_data = catalog.json_data
        items = _extract_catalog_items(json_data)
        if not items:
            return Response(
                {'success': False, 'error': 'No tile items found in catalog JSON'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        errors = []

        for i, item in enumerate(items):
            sku = item.get('sku', '').strip()
            if not sku:
                errors.append({'index': i, 'error': 'Missing sku'})
                continue
            try:
                _, was_created = Tile.objects.get_or_create(
                    sku=sku,
                    defaults={
                        'name': item.get('name') or sku,
                        'dimensions': item.get('dimensions') or '30x30cm',
                        'pieces_per_carton': item.get('pieces_per_carton') or 10,
                        'category': item.get('category') or 'Wall',
                        'description': item.get('description') or '',
                        'brand': item.get('brand') or 'other',
                        'series': item.get('series') or '',
                        'tier': item.get('tier') or 'standard',
                        'tile_type': item.get('tile_type') or '',
                        'finish': item.get('finish') or '',
                        'thickness': item.get('thickness') or '',
                        'coverage_per_box': item.get('coverage_per_box') or '',
                        'use_case': item.get('use_case') or '',
                    },
                )
                if was_created:
                    created += 1
            except Exception as e:
                errors.append({'index': i, 'error': str(e)})

        catalog.processed = True
        catalog.save(update_fields=['processed'])

        return Response({
            'success': True,
            'data': {
                'created': created,
                'skipped': len(items) - created - len(errors),
                'errors': errors,
            },
        })


class InventoryOperationViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, CanPerformInventoryOperations]

    @action(detail=False, methods=['post'])
    def receive(self, request):
        serializer = ReceiveInventorySerializer(data=request.data)
        if serializer.is_valid():
            try:
                inventory, movement = InventoryService.receive_inventory(
                    tile_id=serializer.validated_data['tile_id'],
                    batch_number=serializer.validated_data['batch_number'],
                    production_date=serializer.validated_data['production_date'],
                    supplier=serializer.validated_data['supplier'],
                    cartons=serializer.validated_data['cartons'],
                    loose_pieces=serializer.validated_data['loose_pieces'],
                    location=serializer.validated_data['location'],
                    performed_by=request.user,
                    reference=serializer.validated_data.get('reference', '')
                )
                return Response({
                    'success': True,
                    'data': {
                        'inventory': InventorySerializer(inventory).data,
                        'movement': MovementSerializer(movement).data
                    }
                }, status=status.HTTP_201_CREATED)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def issue_dispatch(self, request):
        serializer = DispatchInventorySerializer(data=request.data)
        if serializer.is_valid():
            try:
                inventory, movement = InventoryService.dispatch_inventory(
                    tile_id=serializer.validated_data['tile_id'],
                    batch_id=serializer.validated_data['batch_id'],
                    cartons=serializer.validated_data['cartons'],
                    loose_pieces=serializer.validated_data['loose_pieces'],
                    location=serializer.validated_data['location'],
                    performed_by=request.user,
                    reference=serializer.validated_data.get('reference', '')
                )
                return Response({
                    'success': True,
                    'data': {
                        'inventory': InventorySerializer(inventory).data,
                        'movement': MovementSerializer(movement).data
                    }
                }, status=status.HTTP_200_OK)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def adjust(self, request):
        serializer = AdjustInventorySerializer(data=request.data)
        if serializer.is_valid():
            try:
                inventory, movement = InventoryService.adjust_inventory(
                    tile_id=serializer.validated_data['tile_id'],
                    batch_id=serializer.validated_data['batch_id'],
                    location=serializer.validated_data['location'],
                    new_cartons=serializer.validated_data['new_cartons'],
                    new_loose_pieces=serializer.validated_data['new_loose_pieces'],
                    reason=serializer.validated_data['reason'],
                    performed_by=request.user
                )
                return Response({
                    'success': True,
                    'data': {
                        'inventory': InventorySerializer(inventory).data,
                        'movement': MovementSerializer(movement).data
                    }
                }, status=status.HTTP_200_OK)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def transfer(self, request):
        serializer = TransferInventorySerializer(data=request.data)
        if serializer.is_valid():
            try:
                source_inventory, dest_inventory, movement = InventoryService.transfer_inventory(
                    tile_id=serializer.validated_data['tile_id'],
                    batch_id=serializer.validated_data['batch_id'],
                    from_location=serializer.validated_data['from_location'],
                    to_location=serializer.validated_data['to_location'],
                    cartons=serializer.validated_data['cartons'],
                    loose_pieces=serializer.validated_data['loose_pieces'],
                    performed_by=request.user,
                    reference=serializer.validated_data.get('reference', '')
                )
                return Response({
                    'success': True,
                    'data': {
                        'source_inventory': InventorySerializer(source_inventory).data,
                        'destination_inventory': InventorySerializer(dest_inventory).data,
                        'movement': MovementSerializer(movement).data
                    }
                }, status=status.HTTP_200_OK)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def stock_take(self, request):
        serializer = StockTakeSerializer(data=request.data)
        if serializer.is_valid():
            try:
                result = InventoryService.stock_take(
                    data=serializer.validated_data['data'],
                    performed_by=request.user,
                )
                return Response({'success': True, 'data': result}, status=status.HTTP_201_CREATED)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


User = get_user_model()


class ReportViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsInventoryViewer]

    @action(detail=False, methods=['get'])
    def stock_summary(self, request):
        total_tiles = Tile.objects.count()
        agg = Inventory.objects.aggregate(
            total_cartons=Sum('cartons'),
            total_loose=Sum('loose_pieces'),
        )
        total_cartons = agg['total_cartons'] or 0
        total_loose = agg['total_loose'] or 0
        total_pieces = Inventory.objects.annotate(
            item_total=F('cartons') * F('tile__pieces_per_carton') + F('loose_pieces')
        ).aggregate(total=Sum('item_total'))['total'] or 0
        low_stock_count = Inventory.objects.annotate(
            item_total=F('cartons') * F('tile__pieces_per_carton') + F('loose_pieces')
        ).filter(item_total__lte=50).count()
        location_count = Inventory.objects.values('location').distinct().count()
        total_batches = Batch.objects.count()
        return Response({
            'total_tiles': total_tiles,
            'total_cartons': total_cartons,
            'total_loose_pieces': total_loose,
            'total_pieces': total_pieces,
            'low_stock_count': low_stock_count,
            'location_count': location_count,
            'total_batches': total_batches,
        })

    @action(detail=False, methods=['get'])
    def movement_summary(self, request):
        period = request.query_params.get('period', 'day')
        from django.utils import timezone
        from datetime import timedelta

        now = timezone.now()
        if period == 'week':
            since = now - timedelta(days=30)
            trunc = 'week'
        elif period == 'month':
            since = now - timedelta(days=365)
            trunc = 'month'
        else:
            since = now - timedelta(days=7)
            trunc = 'day'

        from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
        trunc_func = {'day': TruncDay, 'week': TruncWeek, 'month': TruncMonth}[trunc]

        movements = (
            Movement.objects.filter(created_at__gte=since)
            .annotate(period=trunc_func('created_at'))
            .values('period', 'movement_type')
            .annotate(count=Count('id'))
            .order_by('period', 'movement_type')
        )

        by_type = Movement.objects.filter(created_at__gte=since).values('movement_type').annotate(
            count=Count('id'), total_pieces=Sum('cartons_change')
        ).order_by('-count')

        return Response({
            'period': period,
            'since': since.isoformat(),
            'movements': list(movements),
            'by_type': list(by_type),
        })

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from django.http import HttpResponse

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Inventory Stock Summary", styles['Title']))
        elements.append(Spacer(1, 12))

        total_tiles = Tile.objects.count()
        inventory_items = Inventory.objects.select_related('tile').all()
        total_cartons = sum(item.cartons for item in inventory_items)
        total_loose = sum(item.loose_pieces for item in inventory_items)
        total_pieces = sum(item.total_pieces for item in inventory_items)
        low_stock_count = sum(1 for item in inventory_items if item.total_pieces <= 50)

        summary_data = [
            ['Metric', 'Value'],
            ['Total Tiles', str(total_tiles)],
            ['Total Cartons', str(total_cartons)],
            ['Total Loose Pieces', str(total_loose)],
            ['Total Pieces', str(total_pieces)],
            ['Low Stock Items', str(low_stock_count)],
        ]
        summary_table = Table(summary_data, colWidths=[200, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 24))

        elements.append(Paragraph("Low Stock Items (≤ 50 pieces)", styles['Heading2']))
        elements.append(Spacer(1, 8))

        low_items = [item for item in inventory_items if item.total_pieces <= 50]
        low_sorted = sorted(low_items, key=lambda x: x.total_pieces)

        if low_sorted:
            low_data = [['SKU', 'Name', 'Batch', 'Location', 'Total Pieces']]
            for item in low_sorted:
                low_data.append([
                    item.tile.sku, item.tile.name, item.batch.batch_number,
                    item.location, str(item.total_pieces),
                ])
            low_table = Table(low_data, colWidths=[80, 120, 80, 80, 80])
            low_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CC0000')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFE0E0')]),
            ]))
            elements.append(low_table)
        else:
            elements.append(Paragraph("No low stock items.", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="stock_summary.pdf"'
        return response


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]

    @action(detail=True, methods=['post'])
    def set_role(self, request, pk=None):
        user = self.get_object()
        serializer = SetRoleSerializer(data=request.data)
        if serializer.is_valid():
            role = serializer.validated_data['role']
            group = None
            try:
                from django.contrib.auth.models import Group
                group = Group.objects.get(name='inventory_managers')
            except Group.DoesNotExist:
                pass
            if role == 'admin':
                user.is_superuser = True
                user.is_staff = True
                if group:
                    user.groups.add(group)
            elif role == 'manager':
                user.is_superuser = False
                user.is_staff = False
                if group:
                    user.groups.add(group)
            else:
                user.is_superuser = False
                user.is_staff = False
                if group:
                    user.groups.remove(group)
            user.save()
            return Response({'success': True, 'data': UserSerializer(user).data})
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    search_fields = ['name', 'email', 'phone']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']


class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    search_fields = ['name', 'email', 'phone']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']


class SalesOrderViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SalesOrder.objects.select_related('customer', 'created_by').prefetch_related('line_items__tile', 'line_items__batch').all()
    serializer_class = SalesOrderSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['status', 'customer']
    search_fields = ['order_number', 'customer__name']
    ordering_fields = ['order_date', 'order_number']
    ordering = ['-order_date']

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sales_orders.csv"'
        writer = csv.writer(response)
        writer.writerow(['Order #', 'Customer', 'Status', 'Date', 'Total', 'Notes'])
        for order in self.get_queryset():
            writer.writerow([
                order.order_number, order.customer.name, order.status,
                order.order_date.isoformat(), str(order.total_amount), order.notes,
            ])
        return response


class PurchaseOrderViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PurchaseOrder.objects.select_related('supplier', 'created_by').prefetch_related('line_items__tile', 'line_items__batch').all()
    serializer_class = PurchaseOrderSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]
    filterset_fields = ['status', 'supplier']
    search_fields = ['order_number', 'supplier__name']
    ordering_fields = ['order_date', 'order_number']
    ordering = ['-order_date']

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="purchase_orders.csv"'
        writer = csv.writer(response)
        writer.writerow(['Order #', 'Supplier', 'Status', 'Date', 'Expected', 'Notes'])
        for order in self.get_queryset():
            writer.writerow([
                order.order_number, order.supplier.name, order.status,
                order.order_date.isoformat(),
                order.expected_date.isoformat() if order.expected_date else '',
                order.notes,
            ])
        return response


class OrderOperationViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, CanPerformInventoryOperations]

    @action(detail=False, methods=['post'])
    def create_sales_order(self, request):
        serializer = CreateSalesOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = OrderService.create_sales_order(
                    customer_id=serializer.validated_data['customer_id'],
                    items=serializer.validated_data['items'],
                    performed_by=request.user,
                    notes=serializer.validated_data.get('notes', ''),
                )
                return Response({'success': True, 'data': SalesOrderSerializer(order).data}, status=status.HTTP_201_CREATED)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def create_purchase_order(self, request):
        serializer = CreatePurchaseOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = OrderService.create_purchase_order(
                    supplier_id=serializer.validated_data['supplier_id'],
                    items=serializer.validated_data['items'],
                    performed_by=request.user,
                    expected_date=serializer.validated_data.get('expected_date'),
                    notes=serializer.validated_data.get('notes', ''),
                )
                return Response({'success': True, 'data': PurchaseOrderSerializer(order).data}, status=status.HTTP_201_CREATED)
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def confirm_sales_order(self, request):
        serializer = ConfirmOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = OrderService.confirm_sales_order(
                    order_id=serializer.validated_data['order_id'],
                    performed_by=request.user,
                )
                return Response({'success': True, 'data': SalesOrderSerializer(order).data})
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def ship_sales_order(self, request):
        serializer = ConfirmOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = OrderService.ship_sales_order(
                    order_id=serializer.validated_data['order_id'],
                    performed_by=request.user,
                )
                return Response({'success': True, 'data': SalesOrderSerializer(order).data})
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def cancel_sales_order(self, request):
        serializer = ConfirmOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = OrderService.cancel_sales_order(
                    order_id=serializer.validated_data['order_id'],
                    performed_by=request.user,
                )
                return Response({'success': True, 'data': SalesOrderSerializer(order).data})
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def confirm_purchase_order(self, request):
        serializer = ConfirmOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                order = OrderService.confirm_purchase_order(
                    order_id=serializer.validated_data['order_id'],
                    performed_by=request.user,
                )
                return Response({'success': True, 'data': PurchaseOrderSerializer(order).data})
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def receive_purchase_order(self, request):
        serializer = ConfirmOrderSerializer(data=request.data)
        if serializer.is_valid():
            try:
                location = request.data.get('location', 'RECEIVING')
                order = OrderService.receive_purchase_order(
                    order_id=serializer.validated_data['order_id'],
                    performed_by=request.user,
                    location=location,
                )
                return Response({'success': True, 'data': PurchaseOrderSerializer(order).data})
            except (ValidationError, ObjectDoesNotExist) as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated, IsInventoryViewer]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return Notification.objects.all()
        return Notification.objects.filter(user=user) | Notification.objects.filter(user__isnull=True)

    @action(detail=False, methods=['post'])
    def mark_read(self, request):
        ids = request.data.get('ids', [])
        if ids:
            Notification.objects.filter(id__in=ids).update(is_read=True)
        return Response({'success': True})

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        qs = self.get_queryset().filter(is_read=False)
        count = qs.update(is_read=True)
        return Response({'success': True, 'marked_read': count})

    @action(detail=False, methods=['post'])
    def clear_read(self, request):
        qs = self.get_queryset().filter(is_read=True)
        count, _ = qs.delete()
        return Response({'success': True, 'deleted': count})


class SyncConflictViewSet(viewsets.ModelViewSet):
    queryset = SyncConflict.objects.all()
    serializer_class = SyncConflictSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    filterset_fields = ['model_name', 'resolved']
    ordering = ['-created_at']

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        conflict = self.get_object()
        resolution = request.data.get('resolution')
        if resolution not in ('local', 'remote'):
            return Response({'error': "resolution must be 'local' or 'remote'"}, status=status.HTTP_400_BAD_REQUEST)
        conflict.resolution = resolution
        conflict.resolved = True
        conflict.resolved_at = timezone.now()
        conflict.save()

        from django.apps import apps
        model = apps.get_model('inventory', conflict.model_name)
        if resolution == 'remote' and model:
            try:
                instance = model.objects.get(id=conflict.record_id)
                for key, val in conflict.remote_data.items():
                    if key != 'id' and hasattr(instance, key):
                        setattr(instance, key, val)
                instance.save()
            except model.DoesNotExist:
                model.objects.create(**conflict.remote_data)

        return Response(SyncConflictSerializer(conflict).data)


class BarcodeViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsInventoryViewer]

    @action(detail=False, methods=['get'])
    def generate_qr(self, request):
        batch_id = request.query_params.get('batch_id')
        if not batch_id:
            return Response({'error': 'batch_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            batch = Batch.objects.select_related('tile').get(id=batch_id)
        except Batch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
        import qrcode
        from io import BytesIO
        qr_data = f"INV-BATCH:{batch.batch_number}:{batch.tile.sku}"
        qr = qrcode.make(qr_data)
        buffer = BytesIO()
        qr.save(buffer, format='PNG')
        buffer.seek(0)
        from django.http import HttpResponse
        response = HttpResponse(buffer, content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="batch_{batch.batch_number}.png"'
        return response

    @action(detail=False, methods=['get'])
    def lookup(self, request):
        qr_data = request.query_params.get('data', '')
        if not qr_data:
            return Response({'error': 'data parameter required'}, status=status.HTTP_400_BAD_REQUEST)
        if qr_data.startswith('INV-BATCH:'):
            parts = qr_data.split(':')
            if len(parts) >= 3:
                batch_number = parts[1]
                try:
                    batch = Batch.objects.select_related('tile').get(batch_number=batch_number)
                    inventories = Inventory.objects.filter(batch=batch).select_related('tile')
                    inv_serializer = InventorySerializer(inventories, many=True)
                    return Response({
                        'type': 'batch',
                        'batch': BatchSerializer(batch).data,
                        'inventory': inv_serializer.data,
                    })
                except Batch.DoesNotExist:
                    return Response({'error': 'Batch not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'error': 'Unknown barcode format'}, status=status.HTTP_400_BAD_REQUEST)


IMPORT_ORDER = [
    'Group', 'User',
    'Tile', 'Batch', 'Customer', 'Supplier',
    'Inventory', 'Movement', 'AuditLog',
    'SalesOrder', 'PurchaseOrder', 'OrderLineItem',
    'TileCatalog', 'Notification', 'SyncState', 'SyncConflict',
]


def _run_import(export: dict, dry_run: bool = False) -> dict:
    from io import StringIO
    from django.core.serializers import deserialize
    from django.contrib.auth import get_user_model
    from django.contrib.auth.models import Group, Permission

    User = get_user_model()
    counts = {}

    for label in IMPORT_ORDER:
        records = export.get('models', {}).get(label, [])
        if not records:
            counts[label] = {'created': 0, 'skipped': 0}
            continue

        created = 0
        skipped = 0

        if label == 'Group':
            for g in records:
                if dry_run:
                    created += 1
                    continue
                group, was = Group.objects.get_or_create(
                    name=g['name'],
                    defaults={'id': g['id']},
                )
                perm_codenames = g.get('permissions', [])
                if perm_codenames:
                    perms = Permission.objects.filter(codename__in=perm_codenames)
                    group.permissions.add(*perms)
                if was:
                    created += 1
                else:
                    skipped += 1

        elif label == 'User':
            for u in records:
                if dry_run:
                    created += 1
                    continue
                user, was = User.objects.get_or_create(
                    username=u['username'],
                    defaults={
                        'email': u.get('email', ''),
                        'is_superuser': u['is_superuser'],
                        'is_staff': u['is_staff'],
                        'is_active': u.get('is_active', True),
                    },
                )
                if was:
                    user.set_unusable_password()
                    user.save(update_fields=['password'])
                    created += 1
                else:
                    skipped += 1
                group_names = u.get('groups', [])
                if group_names:
                    groups = Group.objects.filter(name__in=group_names)
                    user.groups.add(*groups)

        else:
            model_name = f'inventory.{label}'
            for obj in records:
                if dry_run:
                    created += 1
                    continue
                stream = StringIO(json.dumps([obj]))
                for deserialized_obj in deserialize('json', stream):
                    try:
                        deserialized_obj.save()
                        created += 1
                    except Exception:
                        skipped += 1

        counts[label] = {'created': created, 'skipped': skipped}

    total_created = sum(c['created'] for c in counts.values())
    total_skipped = sum(c['skipped'] for c in counts.values())
    return {'counts': counts, 'total_created': total_created, 'total_skipped': total_skipped}


class AdminExportViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]

    @action(detail=False, methods=['get'])
    def export_data(self, request):
        from io import BytesIO
        from django.http import HttpResponse
        from inventory.management.commands.export_data import MODELS_IN_ORDER
        from django.core.serializers import serialize
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Group
        from datetime import datetime, timezone

        User = get_user_model()
        indent = 2

        export = {
            'version': '1.0',
            'exported_at': datetime.now(timezone.utc).isoformat(),
            'models': {},
        }

        users = User.objects.all().order_by('id')
        export['models']['User'] = [
            {
                'id': str(u.id), 'username': u.username, 'email': u.email,
                'is_superuser': u.is_superuser, 'is_staff': u.is_staff,
                'is_active': u.is_active,
                'groups': list(u.groups.values_list('name', flat=True)),
            }
            for u in users
        ]

        groups = Group.objects.all().order_by('id')
        export['models']['Group'] = [
            {'id': str(g.id), 'name': g.name,
             'permissions': list(g.permissions.values_list('codename', flat=True))}
            for g in groups
        ]

        for label in MODELS_IN_ORDER:
            from django.apps import apps
            model_cls = apps.get_model('inventory', label)
            qs = model_cls.objects.all().order_by('id')
            raw = serialize('json', qs)
            export['models'][label] = json.loads(raw)

        output = json.dumps(export, indent=indent)
        buffer = BytesIO(output.encode())
        response = HttpResponse(buffer, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="inventory_export.json"'
        return response

    @action(detail=False, methods=['post'])
    def import_data(self, request):
        preview = request.data.get('preview', True)
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            export = json.loads(file.read().decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            return Response({'error': f'Invalid JSON: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        result = _run_import(export, dry_run=preview)
        result['version'] = export.get('version', '?')
        result['exported_at'] = export.get('exported_at', '?')
        result['preview'] = preview
        return Response(result)
